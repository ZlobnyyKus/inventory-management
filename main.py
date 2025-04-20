from fastapi import FastAPI, Request, HTTPException, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
import asyncpg
from asyncpg.pool import Pool
from asyncpg import Record
from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List, Dict, Any, Optional, Union
import os
from contextlib import asynccontextmanager
from pydantic_settings import BaseSettings
import json
import logging
import asyncio


# Настройки приложения
class Settings(BaseSettings):
    database_url: str = "postgresql://postgres:kus3110@localhost:5432/mse_db"
    db_host: str = "localhost"
    db_port: int = 5432
    db_name: str = "mse_db"
    db_user: str = "postgres"
    db_password: str = "kus3110"
    db_pool_min: int = 1
    db_pool_max: int = 50
    omo_password: str = "omo123"
    default_password: str = "00000"
    bureau_count: int = 42
    excluded_bureaus: List[int] = [25, 26, 27, 31, 41]
    experts: List[int] = [1, 2, 3, 5, 8, 9]

    class Config:
        env_file = ".env"


settings = Settings()

# Настройка логгирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Глобальные переменные для connection pool
db_pool: Optional[Pool] = None


async def init_default_passwords(conn):
    """Инициализирует пароли по умолчанию"""
    try:
        # Проверяем, есть ли уже записи
        count = await conn.fetchval("SELECT COUNT(*) FROM passwords")
        if count > 0:
            return

        # Добавляем пароль для ОМО
        await conn.execute("""
            INSERT INTO passwords (bureau_number, password)
            VALUES ('omo', $1)
            ON CONFLICT (bureau_number) DO NOTHING
        """, settings.omo_password)

        # Добавляем пароли для бюро
        for i in range(1, settings.bureau_count + 1):
            if i in settings.excluded_bureaus:
                continue
            await conn.execute("""
                INSERT INTO passwords (bureau_number, password)
                VALUES ($1, $2)
                ON CONFLICT (bureau_number) DO NOTHING
            """, f"bureau_{i}", settings.default_password)

        # Добавляем пароли для экспертных составов
        for expert in settings.experts:
            await conn.execute("""
                INSERT INTO passwords (bureau_number, password)
                VALUES ($1, $2)
                ON CONFLICT (bureau_number) DO NOTHING
            """, f"expert_{expert}", settings.default_password)

        logger.info("Инициализированы пароли по умолчанию")
    except Exception as e:
        logger.error(f"Ошибка при инициализации паролей: {str(e)}")
        raise


async def check_and_create_tables():
    """Проверяет и создаёт таблицы при необходимости"""
    try:
        conn = await get_db_connection()
        try:
            # Проверяем существование таблицы passwords
            table_exists = await conn.fetchval("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_name = 'passwords'
                )
            """)

            if not table_exists:
                await conn.execute("""
                    CREATE TABLE passwords (
                        bureau_number VARCHAR(50) PRIMARY KEY,
                        password VARCHAR(100) NOT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                logger.info("Таблица passwords создана")

            # Проверяем существование таблицы records
            table_exists = await conn.fetchval("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_name = 'records'
                )
            """)

            if not table_exists:
                await conn.execute("""
                    CREATE TABLE records (
                        id SERIAL PRIMARY KEY,
                        bureau_number VARCHAR(50) NOT NULL,
                        full_name TEXT,
                        birth_date DATE,
                        snils VARCHAR(20),
                        age_category VARCHAR(20),
                        mse_date DATE,
                        decision_date DATE,
                        reg_date DATE,
                        special_marks TEXT,
                        military_registration VARCHAR(100),
                        document_format VARCHAR(100),
                        purpose VARCHAR(100),
                        mse_type VARCHAR(50),
                        mse_form VARCHAR(50),
                        mse_form_change VARCHAR(50),
                        prev_disability VARCHAR(50),
                        prev_disability_reason VARCHAR(100),
                        prev_disability_term VARCHAR(50),
                        current_disability VARCHAR(50),
                        current_disability_reason VARCHAR(100),
                        current_disability_term VARCHAR(50),
                        main_diagnosis TEXT,
                        pdo_developed VARCHAR(50),
                        procedure_type VARCHAR(50),
                        decision_changed_part VARCHAR(100),
                        tsr_changed VARCHAR(10),
                        sfr_appeal VARCHAR(10),
                        changed VARCHAR(10),
                        ipra_direction VARCHAR(100),
                        ipra_contains_tsr VARCHAR(10),
                        ipra_changes VARCHAR(100),
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                logger.info("Таблица records создана")

            # Заполняем начальные пароли
            await init_default_passwords(conn)

        finally:
            await release_db_connection(conn)
    except Exception as e:
        logger.error(f"Ошибка при проверке/создании таблиц: {str(e)}")
        raise


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Управление жизненным циклом приложения"""
    global db_pool

    # Инициализация connection pool
    db_pool = await asyncpg.create_pool(
        host=settings.db_host,
        port=settings.db_port,
        database=settings.db_name,
        user=settings.db_user,
        password=settings.db_password,
        min_size=settings.db_pool_min,
        max_size=settings.db_pool_max,
        command_timeout=60,
        timeout=30,
        max_inactive_connection_lifetime=300
    )

    # Проверяем и создаём таблицы при необходимости
    try:
        await check_and_create_tables()
    except Exception as e:
        logger.error(f"Ошибка при инициализации БД: {str(e)}")
        raise

    yield

    # Закрытие connection pool при завершении
    if db_pool:
        await db_pool.close()


# Инициализация FastAPI с lifespan
app = FastAPI(title="Система МСЭ", version="1.0.0", lifespan=lifespan)

# Настройка CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Настройка статических файлов
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# Стили для Excel
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
HEADER_FONT = Font(bold=True)
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Словари для преобразования значений
COMMON_VALUE_MAPPINGS = {
    "age_category": {
        "adult": "Взрослые",
        "child": "Дети"
    },
    "special_marks": {
        "amputee": "Ампутант",
        "prisoner": "Лицо, находящееся в местах лишения свободы",
        "pni": "Лицо, находящееся в ПНИ (ДДИ)",
        "other_institution": "Лицо, находящееся в других стационарных учреждениях соцзащиты",
        "palliative": "Нуждающийся в паллиативной помощи",
        "new_region": "Житель новых регионов",
        "svo": "Участник СВО"
    },
    "military_registration": {
        "registered": "Состоящий на воинском учете",
        "obliged": "Не состоящий на воинском учете, но обязанный состоять",
        "applying": "Поступающий на воинский учет",
        "not_registered": "Не состоящий на воинском учете"
    },
    "mse_form": {
        "in_person": "Очно",
        "in_absentia": "Заочно"
    },
    "mse_form_change": {
        "v": "В",
        "g": "Г",
        "d": "Д",
        "e": "Е",
        "zh": "Ж"
    },
    "prev_disability": {
        "first": "Первая",
        "second": "Вторая",
        "third": "Третья",
        "kri": "КРИ",
        "not_set": "Инвалидность не установлена",
        "supt_set": "Установление СУПТ",
        "supt_not_set": "СУПТ не установлена"
    },
    "prev_disability_reason": {
        "general_disease": "Общее заболевание",
        "childhood": "Инвалидность с детства",
        "professional_disease": "Профессиональное заболевание",
        "work_injury": "Трудовое увечье",
        "military_injury": "Военная травма",
        "military_service": "Заболевание получено в период военной службы",
        "other": "Другое"
    },
    "prev_disability_term": {
        "1_year": "1 год",
        "2_years": "2 года",
        "5_years": "5 лет",
        "until_14": "До 14 лет",
        "until_18": "До 18 лет",
        "indefinitely": "Бессрочно"
    },
    "current_disability": {
        "first": "Первая",
        "second": "Вторая",
        "third": "Третья",
        "kri": "КРИ",
        "not_set": "Инвалидность не установлена",
        "supt_set": "Установление СУПТ",
        "supt_not_set": "СУПТ не установлена"
    },
    "current_disability_reason": {
        "general_disease": "Общее заболевание",
        "childhood": "Инвалидность с детства",
        "professional_disease": "Профессиональное заболевание",
        "work_injury": "Трудовое увечье",
        "military_injury": "Военная травма",
        "military_service": "Заболевание получено в период военной службы",
        "other": "Другое"
    },
    "current_disability_term": {
        "1_year": "1 год",
        "2_years": "2 года",
        "5_years": "5 лет",
        "until_14": "До 14 лет",
        "until_18": "До 18 лет",
        "indefinitely": "Бессрочно"
    },
    "pdo_developed": {
        "consent": "Получено согласие",
        "refusal": "Отказ"
    }
}

BUREAU_VALUE_MAPPINGS = {
    "document_format": {
        "paper_direction": "Направление бумажное",
        "electronic_direction": "Направление в электронном виде",
        "paper_application": "Заявление бумажное",
        "epgu": "Заявление ЕПГУ"
    },
    "purpose": {
        "disability_group": "Группа инвалидности",
        "disabled_child": "Категория 'ребенок-инвалид'",
        "disability_reason": "Причина инвалидности",
        "disability_term": "Срок инвалидности",
        "supt": "Определение СУПТ",
        "ipra_development": "Разработка ИПРА",
        "prp_development": "Разработка ПРП",
        "death_reason": "Определение причины смерти",
        "care_need": "Определение нуждаемости в постоянном постороннем уходе",
        "new_certificate": "Выдача новой справки об инвалидности",
        "duplicate_certificate": "Выдача дубликата справки об инвалидности",
        "new_supt_certificate": "Выдача новой справки о СУПТ",
        "duplicate_supt_certificate": "Выдача дубликата справки о СУПТ",
        "ipra_changes": "Внесение изменений в ИПРА",
        "prp_changes": "Внесение изменений в ПРП"
    },
    "mse_type": {
        "primary": "Первично",
        "repeat": "Повторно",
        "early_repeat": "Повторно (досрочно)"
    },
    "ipra_direction": {
        "regular": "В очередной срок",
        "exclusively": "Исключительно для разработки ИПРА",
        "health_change": "Изменение состояния здоровья (повторно досрочно)"
    },
    "ipra_contains_tsr": {
        "yes": "Да",
        "no": "Нет"
    },
    "ipra_changes": {
        "anthropometric": "Антропометрические данные",
        "personal": "Персональные данные",
        "tsr_specs": "Уточнение технических характеристик ТСР",
        "errors": "Исправление технических ошибок",
        "maternal_capital": "Материнский капитал"
    }
}

EXPERT_VALUE_MAPPINGS = {
    "document_format": {
        "paper_application": "Заявление бумажное",
        "epgu": "Заявление ЕПГУ"
    },
    "procedure_type": {
        "appeal": "Обжалование",
        "control": "Контроль",
        "pdo": "ПДО"
    },
    "purpose": {
        "disability_group": "Группа инвалидности",
        "disabled_child": "Категория ребенок-инвалид",
        "disability_reason": "Причина инвалидности",
        "disability_term": "Срок инвалидности",
        "supt": "Определение СУПТ",
        "care_need": "Определение нуждаемости в постороннем уходе",
        "death_reason": "Определение причины смерти",
        "ipra_development": "Разработка ИПРА",
        "prp_development": "Разработка ПРП",
        "other": "Иное"
    },
    "mse_type": {
        "primary": "Первично",
        "repeat": "Повторно",
        "early_repeat": "Повторно (досрочно)"
    },
    "decision_changed_part": {
        "disability_group": "Группы инвалидности",
        "disabled_child": "Категории ребенок-инвалид",
        "disability_term": "Срока инвалидности",
        "disability_reason": "Причины инвалидности",
        "ipra_development": "Разработки ИПРА",
        "supt": "Степени УПТ",
        "prp_development": "ПРП",
        "other": "Другие"
    },
    "tsr_changed": {
        "yes": "Да"
    },
    "sfr_appeal": {
        "yes": "Да"
    },
    "changed": {
        "yes": "Да"
    }
}


async def get_db_connection():
    """Получает соединение из пула"""
    if not db_pool:
        raise HTTPException(status_code=500, detail="Database pool not initialized")
    return await db_pool.acquire()


async def release_db_connection(conn):
    """Освобождает соединение обратно в пул"""
    if db_pool and conn:
        await db_pool.release(conn)


async def initialize_database():
    """Инициализирует базу данных при первом запуске"""
    try:
        conn = await get_db_connection()
        try:
            # Создание таблиц (асинхронная версия)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS passwords (
                    bureau_number VARCHAR(50) PRIMARY KEY,
                    password VARCHAR(100) NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            await conn.execute("""
                CREATE TABLE IF NOT EXISTS records (
                    id SERIAL PRIMARY KEY,
                    bureau_number VARCHAR(50) NOT NULL,
                    full_name TEXT,
                    birth_date DATE,
                    snils VARCHAR(20),
                    age_category VARCHAR(20),
                    mse_date DATE,
                    decision_date DATE,
                    reg_date DATE,
                    special_marks TEXT,
                    military_registration VARCHAR(100),
                    document_format VARCHAR(100),
                    purpose VARCHAR(100),
                    mse_type VARCHAR(50),
                    mse_form VARCHAR(50),
                    mse_form_change VARCHAR(50),
                    prev_disability VARCHAR(50),
                    prev_disability_reason VARCHAR(100),
                    prev_disability_term VARCHAR(50),
                    current_disability VARCHAR(50),
                    current_disability_reason VARCHAR(100),
                    current_disability_term VARCHAR(50),
                    main_diagnosis TEXT,
                    pdo_developed VARCHAR(50),
                    procedure_type VARCHAR(50),
                    decision_changed_part VARCHAR(100),
                    tsr_changed VARCHAR(10),
                    sfr_appeal VARCHAR(10),
                    changed VARCHAR(10),
                    ipra_direction VARCHAR(100),
                    ipra_contains_tsr VARCHAR(10),
                    ipra_changes VARCHAR(100),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Инициализация паролей
            await conn.execute("""
                INSERT INTO passwords (bureau_number, password)
                VALUES ($1, $2)
                ON CONFLICT (bureau_number) DO NOTHING
            """, "omo", settings.omo_password)

            for i in range(1, settings.bureau_count + 1):
                if i in settings.excluded_bureaus:
                    continue
                await conn.execute("""
                    INSERT INTO passwords (bureau_number, password)
                    VALUES ($1, $2)
                    ON CONFLICT (bureau_number) DO NOTHING
                """, f"bureau_{i}", settings.default_password)

            for expert in settings.experts:
                await conn.execute("""
                    INSERT INTO passwords (bureau_number, password)
                    VALUES ($1, $2)
                    ON CONFLICT (bureau_number) DO NOTHING
                """, f"expert_{expert}", settings.default_password)

        finally:
            await release_db_connection(conn)

    except Exception as e:
        logger.error(f"Ошибка инициализации БД: {str(e)}")
        raise HTTPException(status_code=500, detail="Ошибка инициализации БД")


async def get_password(bureau_number: str) -> Optional[str]:
    """Получает пароль для указанного бюро"""
    try:
        conn = await get_db_connection()
        try:
            result = await conn.fetchval("""
                SELECT password FROM passwords 
                WHERE bureau_number = $1
            """, bureau_number)
            return result
        finally:
            await release_db_connection(conn)
    except Exception as e:
        logger.error(f"Ошибка получения пароля: {str(e)}")
        return None


async def get_records(
        bureau_number: Optional[str] = None,
        limit: Optional[int] = None,
        offset: Optional[int] = None,
        search_query: Optional[str] = None,
        order_by: str = "ORDER BY mse_date DESC"
) -> List[Dict[str, Any]]:
    """Получает записи из базы данных"""
    try:
        conn = await get_db_connection()
        try:
            conditions = []
            params = []
            param_count = 1

            if bureau_number and bureau_number.lower() != "omo":
                conditions.append(f"bureau_number = ${param_count}")
                params.append(bureau_number)
                param_count += 1

            if search_query:
                clean_search = clean_snils(search_query)
                conditions.append(f"""
                    (REPLACE(REPLACE(snils, '-', ''), ' ', '') LIKE ${param_count} OR 
                    full_name ILIKE ${param_count + 1})
                """)
                params.extend([f"%{clean_search}%", f"%{search_query}%"])
                param_count += 2

            where_clause = " WHERE " + " AND ".join(conditions) if conditions else ""

            query = f"""
            SELECT 
                id,
                bureau_number,
                full_name as "fullName",
                birth_date as "birthDate",
                snils,
                age_category as "ageCategory",
                TO_CHAR(mse_date::date, 'YYYY-MM-DD') as "mseDate",
                TO_CHAR(decision_date::date, 'YYYY-MM-DD') as "decisionDate",
                TO_CHAR(reg_date::date, 'YYYY-MM-DD') as "regDate",
                special_marks as "specialMarks",
                military_registration as "militaryRegistration",
                document_format as "documentFormat",
                purpose,
                mse_type as "mseType",
                mse_form as "mseForm",
                mse_form_change as "mseFormChange",
                prev_disability as "prevDisability",
                prev_disability_reason as "prevDisabilityReason",
                prev_disability_term as "prevDisabilityTerm",
                current_disability as "currentDisability",
                current_disability_reason as "currentDisabilityReason",
                current_disability_term as "currentDisabilityTerm",
                main_diagnosis as "mainDiagnosis",
                pdo_developed as "pdoDeveloped",
                procedure_type as "procedureType",
                decision_changed_part as "decisionChangedPart",
                tsr_changed as "tsrChanged",
                sfr_appeal as "sfrAppeal",
                changed,
                ipra_direction as "ipraDirection",
                ipra_contains_tsr as "ipraContainsTsr",
                ipra_changes as "ipraChanges"
            FROM records
            {where_clause}
            {order_by}
            """

            if limit is not None:
                query += f" LIMIT ${param_count}"
                params.append(limit)
                param_count += 1
            if offset is not None:
                query += f" OFFSET ${param_count}"
                params.append(offset)
                param_count += 1

            logger.debug(f"Executing query: {query}")
            logger.debug(f"With params: {params}")

            records = await conn.fetch(query, *params)
            return [dict(record) for record in records]
        finally:
            await release_db_connection(conn)
    except Exception as e:
        logger.error(f"Ошибка получения записей: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ошибка при получении записей из базы данных")


async def get_record_by_id(record_id: int) -> Optional[Dict[str, Any]]:
    """Получает запись МСЭ по ID"""
    try:
        conn = await get_db_connection()
        try:
            record = await conn.fetchrow("""
                SELECT * FROM records 
                WHERE id = $1
            """, record_id)
            return dict(record) if record else None
        finally:
            await release_db_connection(conn)
    except Exception as e:
        logger.error(f"Ошибка получения записи по ID: {str(e)}")
        raise HTTPException(status_code=500, detail="Ошибка при получении записи из базы данных")


async def create_or_update_record(bureau_number: str, record_data: Dict[str, Any],
                                  record_id: Optional[int] = None) -> int:
    """Создает или обновляет запись МСЭ"""
    try:
        conn = await get_db_connection()
        try:
            # Преобразуем ключи из camelCase в snake_case и нормализуем названия полей экспертных составов
            transformed_data = {}
            for key, value in record_data.items():
                if key.lower() == "bureanumber":
                    continue

                # Преобразование названий полей экспертных составов в стандартные
                if key == "expertDocumentFormat":
                    key = "documentFormat"
                elif key == "expertPurpose":
                    key = "purpose"
                elif key == "pdoDevelopedExpert":
                    key = "pdoDeveloped"

                snake_key = ''.join(['_' + c.lower() if c.isupper() else c for c in key]).lstrip('_')

                # Обработка полей с датами
                if snake_key in ['mse_date', 'decision_date', 'reg_date', 'birth_date']:
                    if value is None or value == '':
                        transformed_data[snake_key] = None
                    elif hasattr(value, 'strftime'):  # Уже объект date/datetime
                        transformed_data[snake_key] = value
                    elif isinstance(value, str):
                        try:
                            # Пробуем разные форматы дат
                            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
                                try:
                                    transformed_data[snake_key] = datetime.strptime(value, fmt).date()
                                    break
                                except ValueError:
                                    continue
                            else:
                                raise ValueError(f"Неверный формат даты: {value}")
                        except Exception as e:
                            raise HTTPException(
                                status_code=400,
                                detail=f"Неверный формат даты для поля {key}. Ожидается YYYY-MM-DD. Ошибка: {str(e)}"
                            )
                    else:
                        transformed_data[snake_key] = value
                # Обработка special_marks
                elif snake_key == 'special_marks' and isinstance(value, list):
                    transformed_data[snake_key] = ','.join(value) if value else None
                # Обработка остальных полей
                else:
                    transformed_data[snake_key] = value if value not in (None, '') else None

            # Проверка обязательных полей
            if not transformed_data.get('mse_date'):
                raise HTTPException(status_code=400, detail="Обязательное поле 'Дата проведения МСЭ' не заполнено")

            # Нормализация ФИО
            if 'full_name' in transformed_data and transformed_data['full_name']:
                transformed_data['full_name'] = ' '.join(transformed_data['full_name'].split())

            if record_id:
                # Убедимся, что record_id - целое число
                try:
                    record_id = int(record_id)
                except (ValueError, TypeError):
                    raise HTTPException(
                        status_code=400,
                        detail="Неверный формат ID записи. Ожидается целое число"
                    )

                # Проверяем существование записи
                existing_record = await get_record_by_id(record_id)
                if not existing_record:
                    raise HTTPException(status_code=404, detail="Запись не найдена")

                # Обновление существующей записи
                set_clause = []
                params = []
                for key, value in transformed_data.items():
                    set_clause.append(f"{key} = ${len(params) + 1}")
                    params.append(value)

                params.append(record_id)

                query = f"""
                UPDATE records 
                SET {", ".join(set_clause)}, updated_at = CURRENT_TIMESTAMP
                WHERE id = ${len(params)}
                RETURNING id
                """

                result = await conn.fetchval(query, *params)
                if not result:
                    raise HTTPException(status_code=404, detail="Запись не найдена после обновления")
                return result
            else:
                # Создание новой записи
                columns = ["bureau_number"]
                placeholders = ["$1"]
                values = [bureau_number]

                for key, value in transformed_data.items():
                    columns.append(key)
                    placeholders.append(f"${len(values) + 1}")
                    values.append(value)

                # Удаляем id из данных, если он случайно попал
                if 'id' in columns:
                    idx = columns.index('id')
                    del columns[idx]
                    del placeholders[idx]
                    del values[idx]

                query = f"""
                INSERT INTO records ({", ".join(columns)})
                VALUES ({", ".join(placeholders)})
                RETURNING id
                """

                result = await conn.fetchval(query, *values)
                if not result:
                    raise HTTPException(status_code=500, detail="Не удалось создать запись")
                return result
        finally:
            await release_db_connection(conn)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при сохранении записи: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail="Ошибка при сохранении записи в базу данных"
        )


async def delete_record(record_id: int) -> bool:
    """Удаляет запись МСЭ"""
    try:
        conn = await get_db_connection()
        try:
            result = await conn.execute("""
                DELETE FROM records 
                WHERE id = $1
            """, record_id)
            return "DELETE" in result
        finally:
            await release_db_connection(conn)
    except Exception as e:
        logger.error(f"Ошибка при удалении записи: {str(e)}")
        raise HTTPException(status_code=500, detail="Ошибка при удалении записи из базы данных")


async def count_records(
        bureau_number: Optional[str] = None,
        search_query: Optional[str] = None
) -> int:
    """Подсчитывает количество записей"""
    try:
        conn = await get_db_connection()
        try:
            conditions = []
            params = []
            param_count = 1

            if bureau_number and bureau_number.lower() != "omo":
                conditions.append(f"bureau_number = ${param_count}")
                params.append(bureau_number)
                param_count += 1

            if search_query:
                clean_search = clean_snils(search_query)
                conditions.append(f"""
                    (REPLACE(REPLACE(snils, '-', ''), ' ', '') LIKE ${param_count} OR 
                    full_name ILIKE ${param_count + 1})
                """)
                params.extend([f"%{clean_search}%", f"%{search_query}%"])
                param_count += 2

            where_clause = " WHERE " + " AND ".join(conditions) if conditions else ""

            query = f"SELECT COUNT(*) FROM records {where_clause}"

            count = await conn.fetchval(query, *params)
            return count
        finally:
            await release_db_connection(conn)
    except Exception as e:
        logger.error(f"Ошибка при подсчете записей: {str(e)}")
        raise


def clean_snils(snils: str) -> str:
    """Очищает СНИЛС от нецифровых символов"""
    if not snils:
        return ""
    return ''.join(c for c in snils if c.isdigit())


def format_excel_date(date_value: Union[str, date, None]) -> str:
    """Форматирует дату в формат ДД.ММ.ГГГГ"""
    if date_value is None:
        return ""

    # Если уже объект date/datetime
    if hasattr(date_value, 'strftime'):
        try:
            return date_value.strftime("%d.%m.%Y")
        except (AttributeError, ValueError):
            return str(date_value)

    # Если строка
    if isinstance(date_value, str):
        # Пробуем разные форматы
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
            try:
                date_obj = datetime.strptime(date_value, fmt)
                return date_obj.strftime("%d.%m.%Y")
            except ValueError:
                continue

    return str(date_value)


def map_value(bureau_number: str, field_name: str, value: Optional[str]) -> str:
    """Преобразует значение из базы данных в читаемый формат"""
    if not value:
        return ""

    mappings = get_value_mappings(bureau_number)

    if field_name in mappings:
        mapping = mappings[field_name]

        if ',' in value:
            return ', '.join([
                mapping.get(v.strip(), v.strip())
                for v in value.split(',')
                if v.strip()
            ])

        return mapping.get(value, value)

    return value


def get_value_mappings(bureau_number: str) -> dict:
    """Возвращает соответствующие маппинги значений в зависимости от типа бюро"""
    if bureau_number.startswith("expert_"):
        return {**COMMON_VALUE_MAPPINGS, **EXPERT_VALUE_MAPPINGS}
    return {**COMMON_VALUE_MAPPINGS, **BUREAU_VALUE_MAPPINGS}


def create_excel_worksheet(worksheet, records: List[Dict[str, Any]], is_all_records: bool = False):
    """Заполняет лист Excel данными из записей МСЭ"""
    headers = [
        "Источник" if is_all_records else "№ п/п",
        "ФИО",
        "Дата рождения",
        "СНИЛС",
        "Дата проведения МСЭ",
        "Дата вынесения решения",
        "Дата регистрации направления",
        "Возрастная категория",
        "Особые отметки",
        "Воинский учет",
        "Формат документа",
        "Цель освидетельствования",
        "МСЭ проводится",
        "Форма проведения МСЭ",
        "Изменение формы проведения МСЭ",
        "Инвалидность (предыдущая МСЭ)",
        "Причина инвалидности (предыдущая МСЭ)",
        "Срок инвалидности (предыдущая МСЭ)",
        "Инвалидность (текущая МСЭ)",
        "Причина инвалидности (текущая МСЭ)",
        "Срок инвалидности (текущая МСЭ)",
        "Диагноз основной (с шифром МКБ)",
        "ПДО разработана"
    ]

    is_expert = False
    if records:
        first_record = records[0]
        is_expert = first_record.get("bureau_number", "").startswith("expert_")

    if is_expert:
        headers.extend([
            "Порядок проведения МСЭ",
            "Решение изменено в части",
            "Из них изменено в части ТСР",
            "МСЭ по обращению ОСФР",
            "Из них изменено"
        ])
    else:
        headers.extend([
            "Разработана ИПРА по направлению",
            "ИПРА содержит ТСР",
            "Внесение изменений в ИПРА"
        ])

    # Заполнение заголовков
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    if not records:
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=1)
        cell = worksheet.cell(row=2, column=1, value="Нет записей")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

        for col_num in range(2, len(headers) + 1):
            cell = worksheet.cell(row=2, column=col_num, value="")
            cell.border = THIN_BORDER

        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = 15
        return

    for row_num, record in enumerate(records, 2):
        bureau_number = record.get("bureau_number", "")

        source = ""
        if is_all_records:
            if bureau_number.startswith("bureau_"):
                source = f"Бюро №{bureau_number.split('_')[1]}"
            elif bureau_number.startswith("expert_"):
                source = f"ЭС №{bureau_number.split('_')[1]}"
            else:
                source = bureau_number

        worksheet.cell(
            row=row_num, column=1,
            value=source if is_all_records else row_num - 1
        )
        worksheet.cell(
            row=row_num, column=2,
            value=record.get("fullName") or record.get("full_name") or ""
        )
        worksheet.cell(
            row=row_num, column=3,
            value=format_excel_date(record.get("birthDate") or record.get("birth_date"))
        )
        worksheet.cell(
            row=row_num, column=4,
            value=record.get("snils", "")
        )
        worksheet.cell(
            row=row_num, column=5,
            value=format_excel_date(record.get("mseDate") or record.get("mse_date"))
        )
        worksheet.cell(
            row=row_num, column=6,
            value=format_excel_date(record.get("decisionDate") or record.get("decision_date"))
        )
        worksheet.cell(
            row=row_num, column=7,
            value=format_excel_date(record.get("regDate") or record.get("reg_date"))
        )
        worksheet.cell(
            row=row_num, column=8,
            value=map_value(bureau_number, "age_category", record.get("ageCategory") or record.get("age_category"))
        )
        worksheet.cell(
            row=row_num, column=9,
            value="; ".join([
                map_value(bureau_number, "special_marks", mark.strip())
                for mark in (record.get("specialMarks") or record.get("special_marks") or "").split(",")
                if mark.strip()
            ])
        )
        worksheet.cell(
            row=row_num, column=10,
            value=map_value(bureau_number, "military_registration",
                            record.get("militaryRegistration") or record.get("military_registration"))
        )
        worksheet.cell(
            row=row_num, column=11,
            value=map_value(bureau_number, "document_format",
                            record.get("documentFormat") or record.get("document_format"))
        )
        worksheet.cell(
            row=row_num, column=12,
            value=map_value(bureau_number, "purpose", record.get("purpose"))
        )
        worksheet.cell(
            row=row_num, column=13,
            value=map_value(bureau_number, "mse_type", record.get("mseType") or record.get("mse_type"))
        )
        worksheet.cell(
            row=row_num, column=14,
            value=map_value(bureau_number, "mse_form", record.get("mseForm") or record.get("mse_form"))
        )
        worksheet.cell(
            row=row_num, column=15,
            value=map_value(bureau_number, "mse_form_change",
                            record.get("mseFormChange") or record.get("mse_form_change"))
        )
        worksheet.cell(
            row=row_num, column=16,
            value=map_value(bureau_number, "prev_disability",
                            record.get("prevDisability") or record.get("prev_disability"))
        )
        worksheet.cell(
            row=row_num, column=17,
            value=map_value(bureau_number, "prev_disability_reason",
                            record.get("prevDisabilityReason") or record.get("prev_disability_reason"))
        )
        worksheet.cell(
            row=row_num, column=18,
            value=map_value(bureau_number, "prev_disability_term",
                            record.get("prevDisabilityTerm") or record.get("prev_disability_term"))
        )
        worksheet.cell(
            row=row_num, column=19,
            value=map_value(bureau_number, "current_disability",
                            record.get("currentDisability") or record.get("current_disability"))
        )
        worksheet.cell(
            row=row_num, column=20,
            value=map_value(bureau_number, "current_disability_reason",
                            record.get("currentDisabilityReason") or record.get("current_disability_reason"))
        )
        worksheet.cell(
            row=row_num, column=21,
            value=map_value(bureau_number, "current_disability_term",
                            record.get("currentDisabilityTerm") or record.get("current_disability_term"))
        )
        worksheet.cell(
            row=row_num, column=22,
            value=record.get("mainDiagnosis") or record.get("main_diagnosis") or ""
        )
        worksheet.cell(
            row=row_num, column=23,
            value=map_value(bureau_number, "pdo_developed", record.get("pdoDeveloped") or record.get("pdo_developed"))
        )

        if is_expert:
            worksheet.cell(
                row=row_num, column=24,
                value=map_value(bureau_number, "procedure_type",
                                record.get("procedureType") or record.get("procedure_type"))
            )
            worksheet.cell(
                row=row_num, column=25,
                value=map_value(bureau_number, "decision_changed_part",
                                record.get("decisionChangedPart") or record.get("decision_changed_part"))
            )
            worksheet.cell(
                row=row_num, column=26,
                value=map_value(bureau_number, "tsr_changed", record.get("tsrChanged") or record.get("tsr_changed"))
            )
            worksheet.cell(
                row=row_num, column=27,
                value=map_value(bureau_number, "sfr_appeal", record.get("sfrAppeal") or record.get("sfr_appeal"))
            )
            worksheet.cell(
                row=row_num, column=28,
                value=map_value(bureau_number, "changed", record.get("changed"))
            )
        else:
            worksheet.cell(
                row=row_num, column=24,
                value=map_value(bureau_number, "ipra_direction",
                                record.get("ipraDirection") or record.get("ipra_direction"))
            )
            worksheet.cell(
                row=row_num, column=25,
                value=map_value(bureau_number, "ipra_contains_tsr",
                                record.get("ipraContainsTsr") or record.get("ipra_contains_tsr"))
            )
            worksheet.cell(
                row=row_num, column=26,
                value=map_value(bureau_number, "ipra_changes", record.get("ipraChanges") or record.get("ipra_changes"))
            )

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)

    if records:
        max_row = len(records) + 1
        max_col = len(headers)
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
        worksheet.freeze_panes = "A2"


async def create_excel_workbook(records: List[Dict[str, Any]], bureau_number: str) -> BytesIO:
    """Создает книгу Excel с записями МСЭ"""
    workbook = Workbook()

    if len(workbook.sheetnames) > 0:
        workbook.remove(workbook.active)

    if bureau_number == "all":
        ws_all = workbook.create_sheet(title="Все записи")
        create_excel_worksheet(ws_all, records, is_all_records=True)

        for i in range(1, settings.bureau_count + 1):
            if i in settings.excluded_bureaus:
                continue

            bureau_num = f"bureau_{i}"
            sheet_name = f"Бюро №{i}"[:31]
            ws = workbook.create_sheet(title=sheet_name)
            bureau_records = [r for r in records if r["bureau_number"] == bureau_num]
            create_excel_worksheet(ws, bureau_records)

        for expert in settings.experts:
            expert_num = f"expert_{expert}"
            sheet_name = f"ЭС №{expert}"[:31]
            ws = workbook.create_sheet(title=sheet_name)
            expert_records = [r for r in records if r["bureau_number"] == expert_num]
            create_excel_worksheet(ws, expert_records)

    else:
        if bureau_number.startswith("bureau_"):
            sheet_name = f"Бюро №{bureau_number.split('_')[1]}"[:31]
        elif bureau_number.startswith("expert_"):
            sheet_name = f"ЭС №{bureau_number.split('_')[1]}"[:31]
        else:
            sheet_name = "Записи"[:31]

        ws = workbook.create_sheet(title=sheet_name)
        create_excel_worksheet(ws, records)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


# API endpoints
@app.get("/")
async def read_root(request: Request):
    """Главная страница приложения"""
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/api/password/{bureau_number}")
async def get_password_endpoint(bureau_number: str) -> Dict[str, Any]:
    password = await get_password(bureau_number)
    if not password:
        raise HTTPException(status_code=404, detail="Бюро не найдено")
    return {"password": password}


@app.post("/api/password/{bureau_number}")
async def update_password_endpoint(
        bureau_number: str,
        request: Request
) -> Dict[str, Any]:
    try:
        data = await request.json()
        new_password = data.get("password")

        if not new_password:
            raise HTTPException(status_code=400, detail="Не указан пароль")

        if not await get_password(bureau_number):
            raise HTTPException(status_code=404, detail="Бюро не найдено")

        conn = await get_db_connection()
        try:
            await conn.execute("""
                INSERT INTO passwords (bureau_number, password)
                VALUES ($1, $2)
                ON CONFLICT (bureau_number) DO UPDATE 
                SET password = EXCLUDED.password, updated_at = CURRENT_TIMESTAMP
            """, bureau_number, new_password)
            return {"success": True, "message": "Пароль успешно обновлен"}
        finally:
            await release_db_connection(conn)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при обновлении пароля: {str(e)}")
        raise HTTPException(status_code=500, detail="Ошибка сервера при обновлении пароля")


@app.get("/api/records")
async def get_all_records(
        skip: int = 0,
        limit: int = 30,
        search: Optional[str] = None
) -> Dict[str, Any]:
    records = await get_records(
        limit=limit,
        offset=skip,
        search_query=search
    )
    return {"data": records}


@app.get("/api/records/count")
async def count_all_records(
        search: Optional[str] = None
) -> Dict[str, int]:
    count = await count_records(search_query=search)
    return {"count": count}


@app.get("/api/records/{bureau_number}")
async def get_bureau_records(
        bureau_number: str,
        skip: int = 0,
        limit: int = 30,
        search: Optional[str] = None
) -> Dict[str, Any]:
    records = await get_records(
        bureau_number=bureau_number,
        limit=limit,
        offset=skip,
        search_query=search
    )
    return {"data": records}


@app.get("/api/records/{bureau_number}/count")
async def count_bureau_records(
        bureau_number: str,
        search: Optional[str] = None
) -> Dict[str, int]:
    count = await count_records(bureau_number=bureau_number, search_query=search)
    return {"count": count}


@app.get("/api/records/{bureau_number}/{record_id}")
async def get_single_record(
        bureau_number: str,
        record_id: int
) -> Dict[str, Any]:
    record = await get_record_by_id(record_id)
    if not record:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    return record


@app.get("/api/omo/records")
async def get_omo_records(
        skip: int = 0,
        limit: int = 30,
        search: Optional[str] = None
) -> Dict[str, Any]:
    records = await get_records(
        bureau_number=None,
        limit=limit,
        offset=skip,
        search_query=search
    )
    return {"data": records}


@app.get("/api/omo/records/count")
async def count_omo_records(
        search: Optional[str] = None
) -> Dict[str, int]:
    count = await count_records(bureau_number=None, search_query=search)
    return {"count": count}


@app.get("/api/records/omo/{record_id}")
async def get_omo_record(record_id: int) -> Dict[str, Any]:
    record = await get_record_by_id(record_id)
    if not record:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    return record


@app.post("/api/records")
async def create_or_update_record_endpoint(request: Request) -> Dict[str, Any]:
    try:
        data = await request.json()
        bureau_number = data.get("bureauNumber")
        record_data = data.get("record", {})
        record_id = data.get("recordId")

        if not bureau_number:
            raise HTTPException(status_code=400, detail="Не указано бюро")

        if "bureauNumber" in record_data:
            del record_data["bureauNumber"]

        if not record_data.get("mseDate"):
            raise HTTPException(status_code=400, detail="Обязательное поле 'Дата проведения МСЭ' не заполнено")

        if record_data.get("fullName"):
            record_data["fullName"] = ' '.join(record_data["fullName"].split())

        # Проверяем существование записи при обновлении
        if record_id:
            existing_record = await get_record_by_id(record_id)
            if not existing_record:
                raise HTTPException(status_code=404, detail="Запись не найдена")

        saved_id = await create_or_update_record(bureau_number, record_data, record_id)

        # Получаем полную обновленную запись
        updated_record = await get_record_by_id(saved_id)
        if not updated_record:
            raise HTTPException(status_code=404, detail="Запись не найдена после сохранения")

        return {
            "success": True,
            "id": saved_id,
            "record": updated_record,
            "message": "Запись успешно сохранена"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при сохранении записи: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ошибка сервера при сохранении записи")


@app.delete("/api/records/{bureau_number}/{record_id}")
async def delete_record_endpoint(
        bureau_number: str,
        record_id: int
) -> Dict[str, Any]:
    try:
        record = await get_record_by_id(record_id)
        if not record:
            raise HTTPException(status_code=404, detail="Запись не найдена")

        # Для ОМО пропускаем проверку bureau_number
        if bureau_number.lower() != "omo" and record["bureau_number"] != bureau_number:
            raise HTTPException(status_code=404, detail="Запись не найдена")

        success = await delete_record(record_id)
        if not success:
            raise HTTPException(status_code=404, detail="Запись не найдена")

        return {"success": True, "message": "Запись успешно удалена"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при удалении записи: {str(e)}")
        raise HTTPException(status_code=500, detail="Ошибка сервера при удалении записи")


@app.get("/api/export/{bureau_number}")
async def export_records(
        bureau_number: str,
        search: Optional[str] = None
) -> Response:
    try:
        # Получаем записи из БД
        records = await get_records(
            bureau_number=bureau_number if bureau_number != "all" else None,
            search_query=search
        )

        # Преобразуем записи для экспорта
        export_data = []
        for record in records:
            formatted_record = {}
            for key, value in record.items():
                # Особое преобразование для полей с датами
                if key.lower() in ['msedate', 'decisiondate', 'regdate', 'birthdate',
                                   'mse_date', 'decision_date', 'reg_date', 'birth_date']:
                    formatted_record[key] = format_excel_date(value)
                else:
                    formatted_record[key] = value
            export_data.append(formatted_record)

        # Создаем Excel файл
        buffer = await create_excel_workbook(export_data, bureau_number)

        # Формируем имя файла
        if bureau_number == "all":
            filename = f"mse_all_records_{datetime.now().strftime('%Y%m%d')}.xlsx"
        elif bureau_number.startswith("bureau_"):
            filename = f"mse_bureau_{bureau_number.split('_')[1]}_records_{datetime.now().strftime('%Y%m%d')}.xlsx"
        elif bureau_number.startswith("expert_"):
            filename = f"mse_expert_{bureau_number.split('_')[1]}_records_{datetime.now().strftime('%Y%m%d')}.xlsx"
        else:
            filename = f"mse_records_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Ошибка экспорта: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail="Произошла ошибка при формировании отчета. Пожалуйста, попробуйте позже."
        )


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000)
